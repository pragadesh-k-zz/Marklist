import openpyxl as xl
import random


def mark(total,pass_mark,headings,students):
	wb = xl.Workbook()
	
	ws = wb.active
	
	cur_sheet = wb[ws.title]
	
	cur_sheet.cell(1,1,"Name")
	
	if len(headings) > len(students):
		maximum = len(headings)
		minimum = len(students)
	else:
		maximum = len(students)
		minimum = len(headings)
	
	
	iter_length = 1
	for i in range(2,maximum+2):
		if maximum == len(headings):
			cur_sheet.cell(1,i,headings[(i-2)])
			if iter_length <= minimum:
				cur_sheet.cell(i,1,students[(i-2)])
			else:
				pass
		else:
			cur_sheet.cell(i,1,students[(i-2)])
			if iter_length <= minimum:
				cur_sheet.cell(1,i,headings[(i-2)])
			else:
				pass
		iter_length+=1
	cur_sheet.cell(1,cur_sheet.max_column+1,"Total")
		
	for row in range(2,cur_sheet.max_row+1):
		total_mark = 0
		for column in range(2,cur_sheet.max_column):
			mark=cur_sheet.cell(row,column,random.randint(pass_mark,total))
			total_mark += mark.value
		cur_sheet.cell(row,cur_sheet.max_column,total_mark)
	
		
	wb.save("marklist.xlsx")
	
	print("\nCompleted! Please look at the Excel document 'marklist.xlsx'")
	
	
total = int(input("Enter the Total Mark : "))
	
pass_mark = int(input("Enter the Pass Mark : "))

sub_total = input("Enter the total subject : ")
sub_name = []
stu_name = []
sub = input("Enter the subjects name : ")
sub_name.append(sub)
for a in range(int(sub_total)-1):
	sub_name.append(input(" "))
stu = input("Enter the students name : \nType 'q' after entering students name\n\n")	
stu_name.append(stu)
for b in range(100):
	nmae = input()
	if nmae.upper() == "Q":
		break
	else:
		stu_name.append(nmae)	
		
mark(total,pass_mark,sub_name,stu_name)